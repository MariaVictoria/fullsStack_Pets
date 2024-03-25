from flask import Flask, request
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# Ruta para manejar el envío de formularios
@app.route('/submit-form', methods=['POST'])
def submit_form():
    # Recibir los datos del formulario
    form_data = request.form
    # Guardar los datos en el archivo Excel
    save_to_excel(form_data)
    # Responder al usuario
    return '¡Los datos se han guardado correctamente en el archivo Excel!'

def save_to_excel(data):
    # Cargar el libro de Excel existente o crear uno nuevo si no existe
    try:
        wb = load_workbook('formulario_data.xlsx')
    except FileNotFoundError:
        wb = Workbook()
    ws = wb.active
    
    # Obtener la fila en la que se debe agregar los nuevos datos
    next_row = ws.max_row + 1
    
    # Escribir los datos en la fila correspondiente
    for index, key in enumerate(data.keys(), start=1):
        ws.cell(row=next_row, column=index, value=data[key])
    
    # Guardar el libro de Excel
    wb.save('formulario_data.xlsx')

if __name__ == '__main__':
    app.run(debug=True)
